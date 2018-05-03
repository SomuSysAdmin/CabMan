"""
This script converts a Cab's Pickup Nodal Point worksheet into a SQLite DB to be used by the program.

Author: Somenath Sinha

Components Used:
================
    * Data Parsing      - Excel Data is parsed via OpenPyXl
    * Geo-coding        - GPS of location obtained using Google Maps API
    * DB                - Database used is SQLite
"""

from openpyxl import load_workbook
import sqlite3
import googlemaps

"""
Names of Related files
"""
xl_file = "NodalPoints.xlsx"
db_file = "NodalPoints.db"
api_key = "AIzaSyDP1Yy2gzWdrrB38GKPOEGiBj5B4I4sa1U"

"""
DB CONNECTION AND TABLE CREATION
"""
conn = sqlite3.connect(db_file)
cur = conn.cursor()
create_city_table_sql = "CREATE TABLE IF NOT EXISTS city(" \
                        "id INT PRIMARY KEY NOT NULL, " \
                        "name VARCHAR(255) NOT NULL, " \
                        "code VARCHAR(50) NOT NULL" \
                        ");"

# npCount isn't `NOT NULL` since the value has to be calculated and isn't known at the time of record insertion. Note
#  that the record must exist for nodalPoint Table to use the entries as Foreign Key.
create_colony_table_sql = "CREATE TABLE IF NOT EXISTS colony(" \
                          "id INT PRIMARY KEY NOT NULL, " \
                          "name VARCHAR(255) NOT NULL, " \
                          "cityId INT NOT NULL, " \
                          "npCount INT, " \
                          "FOREIGN KEY (cityId) REFERENCES city(id)" \
                          ");"

create_nodalPoint_table_sql = "CREATE TABLE IF NOT EXISTS nodalPoints(" \
                              "id INT PRIMARY KEY NOT NULL, " \
                              "name VARCHAR(255) NOT NULL, " \
                              "cityId INT NOT NULL, " \
                              "colonyId INT NOT NULL, " \
                              "lat FLOAT(3,15) NOT NULL, " \
                              "lng FLOAT(3,15) NOT NULL, " \
                              "FOREIGN KEY (cityId) REFERENCES city(id), " \
                              "FOREIGN KEY (colonyId) REFERENCES colony(id)" \
                              ");"

cur.execute(create_city_table_sql)
cur.execute(create_colony_table_sql)
cur.execute(create_nodalPoint_table_sql)

"""
DATA PARSING
"""
book = load_workbook(xl_file)
gmaps = googlemaps.Client(api_key)

"""
Getting the list of unique cities and colonies (correlated to cities) that the excel file has. 
Also creating records for all nodal points.
"""
cities = []
colony = []
np = []

col_count = 0
np_id = 0
err_count = 0
for i, sh_name in enumerate(book.sheetnames):
    col_count += 1
    sh = book[sh_name]  # Each sheet in the workbook
    city_ins_sql = "INSERT INTO city(id, name, code) VALUES ({}, '{}', '{}');".format(i+1,
                                                                                      str(sh['A2'].value).replace(
                                                                                          "'", "''"), sh_name)
    cur.execute(city_ins_sql)
    print("● Inserted CITY Record: CityID: {}; Name: {}; Code: {}".format(i+1, sh['A2'].value, sh_name))
    # print("with code {}".format(city_ins_sql))

    last_col = ""
    np_count = 1
    for r in range(sh.min_row+1, sh.max_row + 1):   # Added +1 to min_row to ignore headers
        """
        Adding record for colony
        """
        present_col = str(sh['B' + str(r)].value).replace("'", "''")
        np_id += 1

        if present_col == last_col:
            np_count += 1
        else:
            if last_col != "":
                colony_alter_np_sql = "UPDATE colony SET npCount = {} WHERE id = {};".format(np_count, col_count)
                cur.execute(colony_alter_np_sql)
                print("└─Updated COLONY Record: Set NP = {3} for CityID: {2}; ColonyID: {0}; Name: {1}"
                      "".format(col_count, last_col, i + 1, np_count))
                # print("with code {}".format(colony_alter_np_sql))
                col_count += 1
                np_count = 1
            else:
                last_col = present_col

            # Stores in order: colony ID, colony name, city ID and NP Count
            colony_ins_sql = "INSERT INTO colony(id, name, cityId) VALUES ({}, '{}', {});".format(
                col_count, present_col, i + 1)
            try:
                cur.execute(colony_ins_sql)
            finally:
                print("├─Inserted COLONY Record: CityID: {2}; ColonyID: {0}; Name: {1}"
                  "".format(col_count, present_col, i + 1))
                # print("with code {}".format(colony_ins_sql))
            last_col = present_col

        """
        Adding record for NP
        """
        # Stores in order: NodalPointID, NP Name, City ID, Colony ID
        np_name = str(sh['C' + str(r)].value).replace("_NP", "").replace("'", "''")
        location = np_name.replace("''", "'") + ", " + present_col + ", " + str(sh['A2'].value)
        try:
            gcode = gmaps.geocode(location)
            lat = gcode[0]['geometry']['location']['lat']
            lng = gcode[0]['geometry']['location']['lng']
        except IndexError as err:
            err_count += 1
            try:
                with open("GPSError.log", 'a') as errorLog, open("manualIntervention.log", 'a') as manInt:
                    print("Index Error for: {} returned Geocode :{} with message: \n{}".format(location, gcode, err),
                          file=errorLog)
                    print("[{}] - Location: {}".format(err_count, location), file=manInt)
                    continue
            except IOError as ioerr:
                print("Error writing to file:", ioerr)

        np_ins_sql = "INSERT INTO nodalPoints(id, name, cityId, colonyId, lat, lng) VALUES ({}, '{}', {}, {}, {}, {});"\
                     "".format(np_id, np_name, i+1, col_count, lat, lng)
        try:
            cur.execute(np_ins_sql)
            conn.commit()
        finally:
            print("│ ├─Inserted NodalPoint Record: CityID: {2}; ColonyID: {3}; NodalPointID: {0}; Name: {1}; Lat: {4}; "
                  "Lat: {5}".format(np_id, np_name, i+1, col_count, lat, lng))
            # print("with code {}".format(np_ins_sql))

    # For the final record
    colony_alter_np_sql = "UPDATE colony SET npCount = {} WHERE id = {};".format(np_count, col_count)
    cur.execute(colony_alter_np_sql)
    print("└─Updated COLONY Record: Set NP = {3} for CityID: {2}; ColonyID: {0}; Name: {1}"
          "".format(col_count, last_col, i + 1, np_count))
    # print("with code {}".format(colony_alter_np_sql))

conn.commit()
conn.close()
print("Completed with {} errors!".format(err_count))

"""
This script tests the basic components to parse data from Excel using openpyxl.

Author: Somenath Sinha

Components Used:
================
    * Data Parsing    -   Excel Data is parsed via OpenPyXl
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as gcl

book = load_workbook("NodalPoints.xlsx")
print("The number of worksheets is {0}".format(len(book.sheetnames)))
print("Worksheet name(s): {0}".format(book.sheetnames))

count = 0
for sh_name in book.sheetnames:
    sh = book[sh_name]  # Each sheet in the workbook
    for r in range(sh.min_row, sh.max_row+1):
        count += 1
        print(count, ":\t", end="")
        for c in range(sh.min_column, sh.max_column+1):
            C = gcl(c)
            print(sh[C+str(r)].value, end="\t")
        print( )